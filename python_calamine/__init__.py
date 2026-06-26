"""
python_calamine — чистая Python-реализация API python-calamine.

Поддерживает: xlsx, xlsm, xls, ods
Python 3.9+, без Rust/C-расширений.

Зависимости (pip install):
    openpyxl   — для xlsx / xlsm
    xlrd       — для xls (xlrd >= 2.0, только .xls)
    odfpy      — для ods  (опционально, можно без него)

Совместимость с оригинальным API:
    CalamineWorkbook.from_path(path, load_tables=False)
    CalamineWorkbook.from_filelike(filelike, load_tables=False)
    CalamineWorkbook.from_object(path_or_filelike, load_tables=False)
    load_workbook(path_or_filelike, load_tables=False)

    workbook.sheet_names          -> List[str]
    workbook.sheets_metadata      -> List[SheetMetadata]
    workbook.table_names          -> List[str] | None
    workbook.path                 -> str | None
    workbook.get_sheet_by_name(name)  -> CalamineSheet
    workbook.get_sheet_by_index(idx)  -> CalamineSheet
    workbook.get_table_by_name(name)  -> CalamineTable
    workbook.close()
    with CalamineWorkbook... as wb:

    sheet.name                    -> str
    sheet.height / width          -> int
    sheet.total_height / total_width -> int
    sheet.start / end             -> tuple[int,int] | None
    sheet.merged_cell_ranges      -> list | None
    sheet.to_python(skip_empty_area=True, nrows=None)
    sheet.iter_rows()

    table.name / sheet / columns  -> str / str / List[str]
    table.height / width / start / end
    table.to_python()
"""

from __future__ import annotations

import datetime
import enum
import io
import os
import zipfile
import xml.etree.ElementTree as ET
from typing import (
    BinaryIO, Iterator, List, Optional, Tuple, Union,
)

__version__ = "0.7.0"  # совместимо с оригинальным python-calamine 0.7.0

# ---------------------------------------------------------------------------
# Public re-exports — совместимость с оригинальным пакетом
# ---------------------------------------------------------------------------
__all__ = [
    "CalamineError",
    "CalamineSheet",
    "CalamineTable",
    "CalamineWorkbook",
    "PasswordError",
    "SheetMetadata",
    "SheetTypeEnum",
    "SheetVisibleEnum",
    "TableNotFound",
    "TablesNotLoaded",
    "TablesNotSupported",
    "WorkbookClosed",
    "WorksheetNotFound",
    "XmlError",
    "ZipError",
    "load_workbook",
]

# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------

class CalamineError(Exception):
    """Базовый класс ошибок."""


class PasswordError(CalamineError):
    """Файл защищён паролем."""


class WorksheetNotFound(CalamineError):
    """Лист не найден."""


class XmlError(CalamineError):
    """Ошибка XML."""


class ZipError(CalamineError):
    """Ошибка ZIP-архива."""


class WorkbookClosed(CalamineError):
    """Книга уже закрыта."""


class TablesNotLoaded(CalamineError):
    """Таблицы не загружены (нужен load_tables=True)."""


class TablesNotSupported(CalamineError):
    """Формат файла не поддерживает таблицы."""


class TableNotFound(CalamineError):
    """Таблица не найдена."""


# ---------------------------------------------------------------------------
# Enums
# ---------------------------------------------------------------------------

class SheetTypeEnum(enum.Enum):
    WorkSheet = "WorkSheet"
    DialogSheet = "DialogSheet"
    MacroSheet = "MacroSheet"
    ChartSheet = "ChartSheet"
    Vba = "Vba"


class SheetVisibleEnum(enum.Enum):
    Visible = "Visible"
    Hidden = "Hidden"
    VeryHidden = "VeryHidden"


# ---------------------------------------------------------------------------
# SheetMetadata
# ---------------------------------------------------------------------------

class SheetMetadata:
    """Метаданные листа."""

    __slots__ = ("name", "typ", "visible")

    def __init__(
        self,
        name: str,
        typ: SheetTypeEnum = SheetTypeEnum.WorkSheet,
        visible: SheetVisibleEnum = SheetVisibleEnum.Visible,
    ) -> None:
        self.name = name
        self.typ = typ
        self.visible = visible

    def __repr__(self) -> str:
        return (
            f"SheetMetadata(name={self.name!r}, typ={self.typ!r}, "
            f"visible={self.visible!r})"
        )

    def __eq__(self, other: object) -> bool:
        if not isinstance(other, SheetMetadata):
            return NotImplemented
        return (self.name, self.typ, self.visible) == (other.name, other.typ, other.visible)


# ---------------------------------------------------------------------------
# Типы значений ячейки
# ---------------------------------------------------------------------------
CellValue = Union[
    int, float, str, bool,
    datetime.time, datetime.date, datetime.datetime, datetime.timedelta,
]
Row = List[CellValue]


# ---------------------------------------------------------------------------
# CalamineSheet
# ---------------------------------------------------------------------------

class CalamineSheet:
    """Лист книги."""

    def __init__(self, name: str, rows: List[Row]) -> None:
        self._name = name
        # rows — уже список списков без trailing пустых строк
        self._rows: List[Row] = rows

    # --- свойства ---

    @property
    def name(self) -> str:
        return self._name

    @property
    def height(self) -> int:
        """Число строк с данными."""
        return len(self._rows)

    @property
    def width(self) -> int:
        """Число столбцов в самой широкой строке."""
        if not self._rows:
            return 0
        return max(len(r) for r in self._rows)

    @property
    def total_height(self) -> int:
        return max(0, self.height - 1)

    @property
    def total_width(self) -> int:
        return max(0, self.width - 1)

    @property
    def start(self) -> Optional[Tuple[int, int]]:
        if not self._rows:
            return None
        return (0, 0)

    @property
    def end(self) -> Optional[Tuple[int, int]]:
        if not self._rows:
            return None
        return (self.height - 1, self.width - 1)

    @property
    def merged_cell_ranges(self) -> Optional[List]:
        """Диапазоны объединённых ячеек.
        Возвращает список или None для форматов без поддержки.
        """
        return getattr(self, "_merged_cell_ranges", [])

    # --- методы ---

    def to_python(
        self,
        skip_empty_area: bool = True,
        nrows: Optional[int] = None,
    ) -> List[Row]:
        """Данные листа как список списков."""
        rows = self._rows
        if nrows is not None:
            rows = rows[:nrows]
        if not skip_empty_area:
            return [list(r) for r in rows]
        # skip_empty_area=True — стандартное поведение calamine (уже учтено при загрузке)
        return [list(r) for r in rows]

    def iter_rows(self) -> Iterator[Row]:
        """Итератор по строкам."""
        for row in self._rows:
            yield list(row)

    def __repr__(self) -> str:
        return f"CalamineSheet(name={self._name!r}, height={self.height}, width={self.width})"


# ---------------------------------------------------------------------------
# CalamineTable
# ---------------------------------------------------------------------------

class CalamineTable:
    """Excel-таблица (только xlsx)."""

    def __init__(
        self,
        name: str,
        sheet: str,
        columns: List[str],
        rows: List[Row],
        start: Optional[Tuple[int, int]] = None,
        end: Optional[Tuple[int, int]] = None,
    ) -> None:
        self._name = name
        self._sheet = sheet
        self._columns = columns
        self._rows = rows  # без строки заголовка
        self._start = start
        self._end = end

    @property
    def name(self) -> str:
        return self._name

    @property
    def sheet(self) -> str:
        return self._sheet

    @property
    def columns(self) -> List[str]:
        return list(self._columns)

    @property
    def height(self) -> int:
        return len(self._rows)

    @property
    def width(self) -> int:
        return len(self._columns)

    @property
    def start(self) -> Optional[Tuple[int, int]]:
        return self._start

    @property
    def end(self) -> Optional[Tuple[int, int]]:
        return self._end

    def to_python(self) -> List[Row]:
        """Данные таблицы как список списков (без строки заголовков)."""
        return [list(r) for r in self._rows]

    def __repr__(self) -> str:
        return (
            f"CalamineTable(name={self._name!r}, sheet={self._sheet!r}, "
            f"columns={self._columns!r})"
        )


# ---------------------------------------------------------------------------
# Внутренние ридеры (backends)
# ---------------------------------------------------------------------------

def _read_xlsx(
    source: Union[str, os.PathLike, BinaryIO],
    load_tables: bool,
) -> Tuple[List[CalamineSheet], List[SheetMetadata], Optional[List[CalamineTable]]]:
    """Читает xlsx/xlsm через openpyxl."""
    try:
        import openpyxl
    except ImportError as e:
        raise CalamineError(
            "Для чтения xlsx/xlsm установите openpyxl: pip install openpyxl"
        ) from e

    try:
        wb = openpyxl.load_workbook(
            source,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except openpyxl.utils.exceptions.InvalidFileException as e:
        raise ZipError(str(e)) from e
    except Exception as e:
        raise CalamineError(str(e)) from e

    sheets: List[CalamineSheet] = []
    metas: List[SheetMetadata] = []

    for ws in wb.worksheets:
        # Тип и видимость
        state = ws.sheet_state  # "visible" | "hidden" | "veryHidden"
        visible_map = {
            "visible": SheetVisibleEnum.Visible,
            "hidden": SheetVisibleEnum.Hidden,
            "veryHidden": SheetVisibleEnum.VeryHidden,
        }
        visible = visible_map.get(state, SheetVisibleEnum.Visible)

        metas.append(SheetMetadata(ws.title, SheetTypeEnum.WorkSheet, visible))

        rows: List[Row] = []
        for row in ws.iter_rows():
            cell_values: Row = []
            for cell in row:
                v = cell.value
                if v is None:
                    v = ""
                elif isinstance(v, bool):
                    pass  # bool перед int!
                elif isinstance(v, datetime.datetime):
                    # Если нет времени — возвращаем date, как оригинальный calamine
                    if v.hour == 0 and v.minute == 0 and v.second == 0 and v.microsecond == 0:
                        v = v.date()
                elif isinstance(v, int):
                    v = float(v)
                cell_values.append(v)
            # Убираем trailing пустые ячейки в строке
            while cell_values and cell_values[-1] == "":
                cell_values.pop()
            rows.append(cell_values)

        # Убираем trailing пустые строки
        while rows and all(c == "" for c in rows[-1]):
            rows.pop()

        sheet = CalamineSheet(ws.title, rows)
        # Объединённые ячейки (недоступны в read_only режиме)
        ranges = []
        mc_attr = getattr(ws, "merged_cells", None)
        if mc_attr is not None:
            for mc in mc_attr.ranges:
                s = (mc.min_row - 1, mc.min_col - 1)
                e = (mc.max_row - 1, mc.max_col - 1)
                ranges.append((s, e))
        sheet._merged_cell_ranges = ranges
        sheets.append(sheet)

    wb.close()

    # Таблицы (xlsx named tables)
    tables: Optional[List[CalamineTable]] = None
    if load_tables:
        tables = _read_xlsx_tables(source)

    return sheets, metas, tables


def _read_xlsx_tables(
    source: Union[str, os.PathLike, BinaryIO],
) -> List[CalamineTable]:
    """Читает Excel-таблицы из xlsx через openpyxl (не read_only)."""
    import openpyxl

    if hasattr(source, "read"):
        source.seek(0)  # type: ignore[union-attr]

    wb = openpyxl.load_workbook(source, data_only=True)
    result: List[CalamineTable] = []

    for ws in wb.worksheets:
        for tbl in ws._tables.values():
            ref = tbl.ref  # e.g. "A1:D10"
            from openpyxl.utils import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(ref)

            # Заголовки
            header_row = []
            for col_idx in range(min_col, max_col + 1):
                cell = ws.cell(row=min_row, column=col_idx)
                header_row.append(str(cell.value) if cell.value is not None else "")

            # Данные (без строки заголовка)
            data_rows: List[Row] = []
            for row_idx in range(min_row + 1, max_row + 1):
                row: Row = []
                for col_idx in range(min_col, max_col + 1):
                    v = ws.cell(row=row_idx, column=col_idx).value
                    if v is None:
                        v = ""
                    elif isinstance(v, bool):
                        pass
                    elif isinstance(v, int):
                        v = float(v)
                    row.append(v)
                data_rows.append(row)

            start = (min_row - 1, min_col - 1)
            end = (max_row - 1, max_col - 1)
            result.append(
                CalamineTable(
                    name=tbl.displayName,
                    sheet=ws.title,
                    columns=header_row,
                    rows=data_rows,
                    start=start,
                    end=end,
                )
            )

    wb.close()
    return result


def _read_xls(
    source: Union[str, os.PathLike, BinaryIO],
    load_tables: bool,
) -> Tuple[List[CalamineSheet], List[SheetMetadata], Optional[List[CalamineTable]]]:
    """Читает xls через xlrd."""
    try:
        import xlrd
    except ImportError as e:
        raise CalamineError(
            "Для чтения xls установите xlrd: pip install xlrd"
        ) from e

    try:
        if hasattr(source, "read"):
            data = source.read()  # type: ignore[union-attr]
            wb = xlrd.open_workbook(file_contents=data)
        else:
            wb = xlrd.open_workbook(str(source))
    except xlrd.biffh.XLRDError as e:
        msg = str(e)
        if "password" in msg.lower() or "encrypted" in msg.lower():
            raise PasswordError(msg) from e
        raise CalamineError(msg) from e

    sheets: List[CalamineSheet] = []
    metas: List[SheetMetadata] = []

    for idx in range(wb.nsheets):
        ws = wb.sheet_by_index(idx)

        # Видимость
        vis_raw = wb.sheet_visibility()[idx] if hasattr(wb, "sheet_visibility") else 0
        vis_map = {0: SheetVisibleEnum.Visible, 1: SheetVisibleEnum.Hidden, 2: SheetVisibleEnum.VeryHidden}
        visible = vis_map.get(vis_raw, SheetVisibleEnum.Visible)
        metas.append(SheetMetadata(ws.name, SheetTypeEnum.WorkSheet, visible))

        rows: List[Row] = []
        for r in range(ws.nrows):
            row: Row = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                ctype = cell.ctype
                v = cell.value
                if ctype == xlrd.XL_CELL_EMPTY or ctype == xlrd.XL_CELL_BLANK:
                    v = ""
                elif ctype == xlrd.XL_CELL_BOOLEAN:
                    v = bool(v)
                elif ctype == xlrd.XL_CELL_NUMBER:
                    # xlrd хранит все числа как float
                    pass
                elif ctype == xlrd.XL_CELL_DATE:
                    try:
                        dt_tuple = xlrd.xldate_as_tuple(v, wb.datemode)
                        if dt_tuple[:3] == (0, 0, 0):
                            v = datetime.time(*dt_tuple[3:])
                        elif dt_tuple[3:] == (0, 0, 0, 0):
                            v = datetime.date(*dt_tuple[:3])
                        else:
                            v = datetime.datetime(*dt_tuple)
                    except Exception:
                        pass
                elif ctype == xlrd.XL_CELL_ERROR:
                    v = ""
                row.append(v)

            while row and row[-1] == "":
                row.pop()
            rows.append(row)

        while rows and all(c == "" for c in rows[-1]):
            rows.pop()

        sheet = CalamineSheet(ws.name, rows)
        sheet._merged_cell_ranges = [
            ((mc[0], mc[2]), (mc[1] - 1, mc[3] - 1))
            for mc in ws.merged_cells
        ]
        sheets.append(sheet)

    if load_tables:
        # xls не поддерживает именованные таблицы
        return sheets, metas, []

    return sheets, metas, None


# ---------------------------------------------------------------------------
# ODS namespace constants
# ---------------------------------------------------------------------------
_ODS_NS = {
    "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
    "table":  "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "text":   "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
    "number": "urn:oasis:names:tc:opendocument:xmlns:datastyle:1.0",
    "calcext": "urn:org:documentfoundation:names:experimental:calc:xmlns:calcext:1.0",
}

_ODS_OFFICE = "urn:oasis:names:tc:opendocument:xmlns:office:1.0"
_ODS_TABLE  = "urn:oasis:names:tc:opendocument:xmlns:table:1.0"
_ODS_TEXT   = "urn:oasis:names:tc:opendocument:xmlns:text:1.0"


def _ods_cell_value(cell_el: ET.Element) -> CellValue:
    """Преобразует XML-элемент ячейки ODS в Python-значение."""
    vtype = cell_el.get(f"{{{_ODS_OFFICE}}}value-type")
    if vtype is None:
        return ""

    if vtype == "string":
        # Собираем текст из всех <text:p>
        parts = []
        for p in cell_el.findall(f".//{{{_ODS_TEXT}}}p"):
            parts.append("".join(p.itertext()))
        return "\n".join(parts)

    if vtype == "float":
        raw = cell_el.get(f"{{{_ODS_OFFICE}}}value")
        if raw is None:
            return ""
        return float(raw)

    if vtype == "percentage":
        raw = cell_el.get(f"{{{_ODS_OFFICE}}}value")
        return float(raw) if raw is not None else ""

    if vtype == "currency":
        raw = cell_el.get(f"{{{_ODS_OFFICE}}}value")
        return float(raw) if raw is not None else ""

    if vtype == "boolean":
        raw = cell_el.get(f"{{{_ODS_OFFICE}}}boolean-value", "")
        return raw.lower() == "true"

    if vtype == "date":
        raw = cell_el.get(f"{{{_ODS_OFFICE}}}date-value", "")
        if "T" in raw:
            try:
                return datetime.datetime.fromisoformat(raw)
            except ValueError:
                pass
        try:
            return datetime.date.fromisoformat(raw[:10])
        except ValueError:
            return raw

    if vtype == "time":
        raw = cell_el.get(f"{{{_ODS_OFFICE}}}time-value", "")
        # ISO 8601 duration: PT1H30M45S
        return _parse_ods_duration(raw)

    return ""


def _parse_ods_duration(s: str) -> Union[datetime.timedelta, str]:
    """PT12H30M0S → timedelta."""
    import re
    m = re.match(
        r"^PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+(?:\.\d+)?)S)?$", s
    )
    if not m:
        return s
    h = int(m.group(1) or 0)
    mi = int(m.group(2) or 0)
    sec_raw = float(m.group(3) or 0)
    sec = int(sec_raw)
    ms = int((sec_raw - sec) * 1_000_000)
    return datetime.timedelta(hours=h, minutes=mi, seconds=sec, microseconds=ms)


def _read_ods(
    source: Union[str, os.PathLike, BinaryIO],
    load_tables: bool,
) -> Tuple[List[CalamineSheet], List[SheetMetadata], Optional[List[CalamineTable]]]:
    """Читает ods через zipfile + ElementTree (без сторонних зависимостей для ODS)."""
    try:
        if hasattr(source, "read"):
            data = source.read()  # type: ignore[union-attr]
            zf = zipfile.ZipFile(io.BytesIO(data))
        else:
            zf = zipfile.ZipFile(str(source))
    except zipfile.BadZipFile as e:
        raise ZipError(str(e)) from e

    try:
        content = zf.read("content.xml")
    except KeyError as e:
        raise XmlError("content.xml не найден в ODS-архиве") from e
    finally:
        zf.close()

    try:
        root = ET.fromstring(content)
    except ET.ParseError as e:
        raise XmlError(str(e)) from e

    spreadsheet = root.find(
        f".//{{{_ODS_OFFICE}}}spreadsheet"
    )
    if spreadsheet is None:
        raise XmlError("Не найден элемент office:spreadsheet")

    sheets: List[CalamineSheet] = []
    metas: List[SheetMetadata] = []

    for tbl in spreadsheet.findall(f"{{{_ODS_TABLE}}}table"):
        name = tbl.get(f"{{{_ODS_TABLE}}}name", "")

        # Видимость (table:display)
        display = tbl.get(f"{{{_ODS_TABLE}}}display", "true")
        visible = (
            SheetVisibleEnum.Visible if display.lower() == "true"
            else SheetVisibleEnum.Hidden
        )
        metas.append(SheetMetadata(name, SheetTypeEnum.WorkSheet, visible))

        rows: List[Row] = []

        for row_el in tbl.findall(f"{{{_ODS_TABLE}}}table-row"):
            row_repeat = int(
                row_el.get(f"{{{_ODS_TABLE}}}number-rows-repeated", "1")
            )

            row: Row = []
            for cell_el in row_el:
                tag = cell_el.tag
                # table:table-cell или table:covered-table-cell
                is_cell = tag in (
                    f"{{{_ODS_TABLE}}}table-cell",
                    f"{{{_ODS_TABLE}}}covered-table-cell",
                )
                if not is_cell:
                    continue

                col_repeat = int(
                    cell_el.get(f"{{{_ODS_TABLE}}}number-columns-repeated", "1")
                )
                val = _ods_cell_value(cell_el)
                row.extend([val] * col_repeat)

            # Убираем trailing пустые ячейки
            while row and row[-1] == "":
                row.pop()

            # Добавляем row_repeat раз (но не хвостовые пустые строки сейчас)
            for _ in range(row_repeat):
                rows.append(list(row))

        # Убираем trailing пустые строки
        while rows and all(c == "" for c in rows[-1]):
            rows.pop()

        sheets.append(CalamineSheet(name, rows))

    tables: Optional[List[CalamineTable]] = None
    if load_tables:
        # ODS не поддерживает именованные таблицы в стиле Excel
        raise TablesNotSupported("ODS не поддерживает именованные таблицы")

    return sheets, metas, tables


# ---------------------------------------------------------------------------
# CalamineWorkbook
# ---------------------------------------------------------------------------

_SUPPORTED_EXT = {".xlsx", ".xlsm", ".xls", ".ods"}


def _detect_format(source: Union[str, os.PathLike, BinaryIO]) -> str:
    """Определяет формат файла по расширению или сигнатуре."""
    if not hasattr(source, "read"):
        ext = os.path.splitext(str(source))[1].lower()
        if ext in _SUPPORTED_EXT:
            return ext
        raise CalamineError(f"Неподдерживаемое расширение файла: {ext!r}")

    # Для файлоподобных объектов — читаем первые байты
    magic = source.read(8)  # type: ignore[union-attr]
    source.seek(0)           # type: ignore[union-attr]

    # PK (ZIP) → xlsx/xlsm/ods
    if magic[:2] == b"PK":
        # Попробуем открыть как zip и посмотреть на содержимое
        try:
            with zipfile.ZipFile(io.BytesIO(magic + source.read())) as zf:  # type: ignore[union-attr]
                source.seek(0)
                names = zf.namelist()
            if "xl/workbook.xml" in names or any(n.startswith("xl/") for n in names):
                return ".xlsx"
            if "content.xml" in names:
                return ".ods"
        except Exception:
            source.seek(0)  # type: ignore[union-attr]
        return ".xlsx"  # best guess

    # D0 CF (OLE2 compound) → xls
    if magic[:2] == b"\xd0\xcf":
        return ".xls"

    raise CalamineError("Не удалось определить формат файла")


class CalamineWorkbook:
    """Рабочая книга Excel/ODS."""

    def __init__(
        self,
        sheets: List[CalamineSheet],
        metas: List[SheetMetadata],
        tables: Optional[List[CalamineTable]],
        path: Optional[str] = None,
    ) -> None:
        self._sheets = sheets
        self._metas = metas
        self._tables = tables
        self._path = path
        self._closed = False

    # --- свойства ---

    @property
    def path(self) -> Optional[str]:
        return self._path

    @property
    def sheet_names(self) -> List[str]:
        return [m.name for m in self._metas]

    @property
    def sheets_metadata(self) -> List[SheetMetadata]:
        return list(self._metas)

    @property
    def table_names(self) -> Optional[List[str]]:
        if self._tables is None:
            raise TablesNotLoaded(
                "Таблицы не загружены. Используйте load_tables=True."
            )
        return [t.name for t in self._tables]

    # --- методы ---

    def get_sheet_by_name(self, name: str) -> CalamineSheet:
        self._check_open()
        for sheet in self._sheets:
            if sheet.name == name:
                return sheet
        raise WorksheetNotFound(f"Лист {name!r} не найден")

    def get_sheet_by_index(self, index: int) -> CalamineSheet:
        self._check_open()
        try:
            return self._sheets[index]
        except IndexError:
            raise WorksheetNotFound(f"Лист с индексом {index} не найден") from None

    def get_table_by_name(self, name: str) -> CalamineTable:
        self._check_open()
        if self._tables is None:
            raise TablesNotLoaded(
                "Таблицы не загружены. Используйте load_tables=True."
            )
        for tbl in self._tables:
            if tbl.name == name:
                return tbl
        raise TableNotFound(f"Таблица {name!r} не найдена")

    def close(self) -> None:
        if self._closed:
            raise WorkbookClosed("Книга уже закрыта")
        self._closed = True
        self._sheets = []
        self._metas = []
        self._tables = None

    def _check_open(self) -> None:
        if self._closed:
            raise WorkbookClosed("Операция невозможна: книга закрыта")

    # --- context manager ---

    def __enter__(self) -> "CalamineWorkbook":
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        if not self._closed:
            self._closed = True

    def __repr__(self) -> str:
        sheets = ", ".join(self.sheet_names) if not self._closed else "<closed>"
        return f"CalamineWorkbook(path={self._path!r}, sheets=[{sheets}])"

    # --- фабричные методы ---

    @classmethod
    def from_path(
        cls,
        path: Union[str, os.PathLike],
        load_tables: bool = False,
    ) -> "CalamineWorkbook":
        """Открыть файл по пути."""
        path_str = str(path)
        ext = os.path.splitext(path_str)[1].lower()
        sheets, metas, tables = _dispatch_read(path, ext, load_tables)
        return cls(sheets, metas, tables, path=path_str)

    @classmethod
    def from_filelike(
        cls,
        filelike: BinaryIO,
        load_tables: bool = False,
    ) -> "CalamineWorkbook":
        """Открыть из файлоподобного объекта."""
        fmt = _detect_format(filelike)
        sheets, metas, tables = _dispatch_read(filelike, fmt, load_tables)
        return cls(sheets, metas, tables, path=None)

    @classmethod
    def from_object(
        cls,
        path_or_filelike: Union[str, os.PathLike, BinaryIO],
        load_tables: bool = False,
    ) -> "CalamineWorkbook":
        """Открыть из пути или файлоподобного объекта."""
        if hasattr(path_or_filelike, "read"):
            return cls.from_filelike(path_or_filelike, load_tables)  # type: ignore[arg-type]
        return cls.from_path(path_or_filelike, load_tables)  # type: ignore[arg-type]


def _dispatch_read(
    source: Union[str, os.PathLike, BinaryIO],
    fmt: str,
    load_tables: bool,
) -> Tuple[List[CalamineSheet], List[SheetMetadata], Optional[List[CalamineTable]]]:
    if fmt in (".xlsx", ".xlsm"):
        return _read_xlsx(source, load_tables)
    if fmt == ".xls":
        return _read_xls(source, load_tables)
    if fmt == ".ods":
        return _read_ods(source, load_tables)
    raise CalamineError(f"Неподдерживаемый формат: {fmt!r}")


# ---------------------------------------------------------------------------
# Удобная функция (аналог openpyxl.load_workbook)
# ---------------------------------------------------------------------------

def load_workbook(
    path_or_filelike: Union[str, os.PathLike, BinaryIO],
    load_tables: bool = False,
) -> CalamineWorkbook:
    """Открыть книгу из пути или файлоподобного объекта."""
    return CalamineWorkbook.from_object(path_or_filelike, load_tables)
